"""
Sincronizacion mensual de senales de contexto.

Lee location_source_config para descubrir que fuentes tiene configuradas
cada ubicacion. Solo ejecuta handlers para ubicaciones con config activa.

CLI:
  python -m src.data_ingestion.sync_mensual
  python -m src.data_ingestion.sync_mensual --location <uuid>
  python -m src.data_ingestion.sync_mensual --solo metro_madrid
  python -m src.data_ingestion.sync_mensual --force
"""

from __future__ import annotations

import io
import json
import os
import re
import urllib.parse
import urllib.request
import warnings
from collections.abc import Callable
from datetime import date

import requests

from src.data_ingestion._common import (
    ensure_feature_registry,
    get_configured_locations,
    get_source_config,
    is_fresh,
    write_month_uniform,
    write_sync_marker,
)

# ── Handlers privados ─────────────────────────────────────────────────────────


def _handler_metro_madrid(
    loc_uuid: str,
    params: dict,
    max_age_hours: float,
    verbose: bool,
) -> int:
    """Validaciones mensuales de Metro de Madrid por estacion."""
    if is_fresh(loc_uuid, "metro_madrid", max_age_hours):
        return 0

    cfg = get_source_config("metro_madrid", params)
    feature_key_prefix = cfg["feature_key_prefix"]

    estaciones = params.get("estaciones", [])
    url_pattern = params.get("anyo_url")
    if not estaciones or not url_pattern:
        if verbose:
            print(f"  [metro_madrid] {loc_uuid}: sin estaciones o anyo_url en params — omitido")
        return 0

    _MESES_COL = [
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]
    _TIMEOUT = 30

    def _download_excel(year: int) -> bytes | None:
        url = url_pattern.format(year=year)
        try:
            r = requests.get(url, timeout=_TIMEOUT)
            if r.status_code in (404, 403):
                return None
            r.raise_for_status()
            return r.content
        except Exception:
            return None

    def _parse_excel_year(xlsx_bytes: bytes, year: int) -> dict[str, dict[int, int]]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl es necesario: pip install openpyxl") from exc

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

        target_sheet = None
        for sh_name in wb.sheetnames:
            if str(year) in sh_name:
                target_sheet = wb[sh_name]
                break
        if target_sheet is None:
            target_sheet = wb.active

        ws = target_sheet
        header_row: int | None = None
        for r in range(1, min(10, ws.max_row + 1)):
            vals_b_m = [ws.cell(r, c).value for c in range(2, 14)]
            non_null = [v for v in vals_b_m if v is not None]
            if len(non_null) >= 6:
                first = non_null[0]
                if isinstance(first, str) and any(m in str(first).lower() for m in _MESES_COL):
                    header_row = r
                    break
                if isinstance(first, (int, float)) and 1 <= int(first) <= 12:
                    header_row = r
                    break

        if header_row is not None:
            col_mes: dict[int, int] = {}
            for c in range(2, 14):
                v = ws.cell(header_row, c).value
                if v is None:
                    continue
                if isinstance(v, str):
                    for i, mn in enumerate(_MESES_COL, 1):
                        if mn in v.lower():
                            col_mes[i] = c
                            break
                elif isinstance(v, (int, float)) and 1 <= int(v) <= 12:
                    col_mes[int(v)] = c
        else:
            col_mes = {m: m + 1 for m in range(1, 13)}
            header_row = 1

        nombre_to_slug = {e["nombre"].strip().lower(): e["slug"] for e in estaciones}
        result: dict[str, dict[int, int]] = {}

        for r in range(header_row + 1, ws.max_row + 1):
            cell_a = ws.cell(r, 1).value
            if cell_a is None:
                continue
            nombre_celda = str(cell_a).strip().lower()
            for nombre_cfg, slug in nombre_to_slug.items():
                if nombre_cfg in nombre_celda or nombre_celda in nombre_cfg:
                    mes_vals: dict[int, int] = {}
                    for mes, col in col_mes.items():
                        v = ws.cell(r, col).value
                        try:
                            mes_vals[mes] = int(float(str(v).replace(",", "").strip()))
                        except (TypeError, ValueError):
                            pass
                    if mes_vals:
                        result[slug] = mes_vals
                    break

        return result

    def _ensure_registry(slug: str) -> None:
        from src.db.store import get_conn

        fk = f"{feature_key_prefix}{slug}"
        ensure_feature_registry(
            fk,
            "metro_madrid",
            "movilidad",
            f"Validaciones mensuales estacion de metro — {slug.replace('_', ' ').title()}",
        )
        get_conn().execute(
            "INSERT INTO feature_flags (feature_key, location_uuid, status, periodicidad) "
            "VALUES (?,?,'contexto','mensual') "
            "ON CONFLICT (feature_key, location_uuid) DO NOTHING",
            [fk, loc_uuid],
        )

    def _sync_year(year: int) -> int:
        raw = _download_excel(year)
        if raw is None:
            if verbose:
                print(f"  [metro_madrid] {year}: descarga fallida o no disponible")
            return 0
        parsed = _parse_excel_year(raw, year)
        if not parsed:
            if verbose:
                print(f"  [metro_madrid] {year}: ninguna estacion encontrada en el Excel")
            return 0
        total = 0
        for slug, mes_vals in parsed.items():
            _ensure_registry(slug)
            fk = f"{feature_key_prefix}{slug}"
            for mes, validaciones in mes_vals.items():
                n = write_month_uniform(year, mes, validaciones, loc_uuid, fk, verbose)
                if verbose and n > 0:
                    print(
                        f"  [metro_madrid] {slug} {mes:02d}/{year}: "
                        f"{validaciones:,} validaciones → {validaciones/n:.0f}/dia"
                    )
                total += n
        return total

    try:
        today = date.today()
        n = _sync_year(today.year - 1)
        n += _sync_year(today.year)
        write_sync_marker(loc_uuid, "metro_madrid")
        return n
    except Exception as e:
        if verbose:
            print(f"  [metro_madrid] {loc_uuid} ERROR — {e}")
        return 0


def _handler_puertos_estado(
    loc_uuid: str,
    params: dict,
    max_age_hours: float,
    verbose: bool,
) -> int:
    """Pasajeros de crucero oficiales — Puertos del Estado."""
    if is_fresh(loc_uuid, "puertos_estado", max_age_hours):
        if verbose:
            port = params.get("port_authority", loc_uuid)
            print(f"  [puertos_estado] {port} omitido (datos < {max_age_hours:.0f}h)")
        return 0

    port_authority = params.get("port_authority")
    if not port_authority:
        if verbose:
            print(f"  [puertos_estado] {loc_uuid}: sin port_authority en params — omitido")
        return 0

    cfg = get_source_config("puertos_estado", params)
    _FEATURE_KEY = cfg["feature_key"]
    _LISTING_URL = cfg["listing_url"]
    _BASE_URL = _LISTING_URL.rsplit("/en/", 1)[0]
    _SHEET = cfg.get("hoja_excel", "Pasajeros crucero")
    _TIMEOUT = 30

    def _fetch_listing_ids(year: int) -> dict[int, int]:
        date_value = 2027 - year
        if date_value < 1:
            return {}
        try:
            r = requests.get(_LISTING_URL, params={"date_value": date_value}, timeout=_TIMEOUT)
            r.raise_for_status()
        except Exception:
            return {}
        all_ids = re.findall(r"/file-download/download/public/(\d+)", r.text)
        pdf_months = re.findall(r"CuadrosResumen_\d{4}_(\d{2})\.pdf", r.text)
        xlsx_ids = all_ids[::2]
        month_ids: dict[int, int] = {}
        for fid, mo_str in zip(xlsx_ids, pdf_months):
            mo = int(mo_str)
            if mo not in month_ids:
                month_ids[mo] = int(fid)
        return month_ids

    def _download_xlsx(file_id: int) -> bytes | None:
        url = f"{_BASE_URL}/file-download/download/public/{file_id}"
        try:
            r = requests.get(url, timeout=_TIMEOUT)
            if r.status_code in (404, 403):
                return None
            r.raise_for_status()
            return r.content
        except Exception:
            return None

    def _parse_xlsx(xlsx_bytes: bytes) -> dict | None:
        _MESES_ES = {
            "enero": 1,
            "febrero": 2,
            "marzo": 3,
            "abril": 4,
            "mayo": 5,
            "junio": 6,
            "julio": 7,
            "agosto": 8,
            "septiembre": 9,
            "octubre": 10,
            "noviembre": 11,
            "diciembre": 12,
        }
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl es necesario: pip install openpyxl") from exc

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

        if _SHEET not in wb.sheetnames:
            return None
        ws = wb[_SHEET]

        month_num: int | None = None
        for col in range(1, ws.max_column + 1):
            v = ws.cell(5, col).value
            if v and isinstance(v, str):
                for name, num in _MESES_ES.items():
                    if name in v.lower():
                        month_num = num
                        break
            if month_num:
                break

        year_prev: int | None = None
        year_act: int | None = None
        for col in range(2, 6):
            v = ws.cell(6, col).value
            if v and isinstance(v, (int, float)):
                yr = int(v)
                if 2010 < yr < 2100:
                    if year_prev is None:
                        year_prev = yr
                    elif year_act is None:
                        year_act = yr
                        break

        if month_num is None or year_act is None or year_prev is None:
            return None

        ap_row: int | None = None
        for r in range(7, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and port_authority.lower() == str(v).strip().lower():
                ap_row = r
                break

        if ap_row is None:
            return None

        def _int(v) -> int:
            try:
                return int(float(str(v).replace(",", "").strip()))
            except (TypeError, ValueError):
                return 0

        return {
            "month": month_num,
            "year_act": year_act,
            "year_prev": year_prev,
            "pax_act": _int(ws.cell(ap_row, 3).value),
            "pax_prev": _int(ws.cell(ap_row, 2).value),
        }

    def _sync_year(year: int) -> int:
        ids = _fetch_listing_ids(year)
        if not ids:
            if verbose:
                print(f"  [puertos_estado] {port_authority} {year}: no se encontraron ficheros")
            return 0
        total = 0
        for mes, fid in sorted(ids.items()):
            raw = _download_xlsx(fid)
            if raw is None:
                if verbose:
                    print(
                        f"  [puertos_estado] {port_authority} {mes:02d}/{year}: "
                        f"descarga fallida (ID {fid})"
                    )
                continue
            parsed = _parse_xlsx(raw)
            if parsed is None:
                if verbose:
                    print(
                        f"  [puertos_estado] {port_authority} {mes:02d}/{year}: "
                        f"formato inesperado o AP no encontrada (ID {fid})"
                    )
                continue
            n = write_month_uniform(
                parsed["year_act"],
                parsed["month"],
                parsed["pax_act"],
                loc_uuid,
                _FEATURE_KEY,
            )
            if n > 0:
                total += n
                if verbose:
                    print(
                        f"  [puertos_estado] {port_authority}"
                        f" {parsed['month']:02d}/{parsed['year_act']}: {parsed['pax_act']:,} pax"
                    )
            n_prev = write_month_uniform(
                parsed["year_prev"],
                parsed["month"],
                parsed["pax_prev"],
                loc_uuid,
                _FEATURE_KEY,
            )
            if n_prev > 0:
                total += n_prev
                if verbose:
                    print(
                        f"  [puertos_estado] {port_authority}"
                        f" {parsed['month']:02d}/{parsed['year_prev']}:"
                        f" {parsed['pax_prev']:,} pax (anyo ant.)"
                    )
        return total

    try:
        today = date.today()
        n = _sync_year(today.year - 1)
        n += _sync_year(today.year)
        write_sync_marker(loc_uuid, "puertos_estado")
        if verbose:
            print(f"  [puertos_estado] {port_authority}: {n} dias escritos")
        return n
    except Exception as e:
        if verbose:
            print(f"  [puertos_estado] {port_authority} ERROR — {e}")
        return 0


def _handler_ine_eoh(
    loc_uuid: str,
    params: dict,
    max_age_hours: float,
    verbose: bool,
) -> int:
    """INE Encuesta de Ocupacion Hotelera — viajeros y pernoctaciones."""
    if is_fresh(loc_uuid, "ine_eoh", max_age_hours):
        return 0

    provincia = params.get("provincia_nombre")
    if not provincia:
        if verbose:
            print(f"  [ine_eoh] {loc_uuid}: sin provincia_nombre en params — omitido")
        return 0

    cfg = get_source_config("ine_eoh", params)
    base_url = cfg["base_url"]
    tabla_viajeros_default = cfg["tabla_viajeros"]
    feature_key_viajeros = cfg["feature_key_viajeros"]
    feature_key_pernoctaciones = cfg["feature_key_pernoctaciones"]
    _TIMEOUT = 30

    def _fetch_tabla(tabla_id: int, nult: int = 300) -> list[dict]:
        url = f"{base_url}/DATOS_TABLA/{tabla_id}?nult={nult}"
        r = requests.get(url, timeout=_TIMEOUT)
        r.raise_for_status()
        return r.json()

    def _parse_serie_mensual(serie: dict) -> list[tuple[int, int, float]]:
        rows = []
        for dp in serie.get("Data", []):
            if dp.get("Secreto") or dp.get("Valor") is None:
                continue
            try:
                year = int(dp["Anyo"])
                periodo = dp.get("FK_Periodo") or dp.get("Periodo", "")
                if isinstance(periodo, int):
                    mes = periodo
                else:
                    mes = int(str(periodo).replace("M", ""))
                if not (1 <= mes <= 12):
                    continue
                rows.append((year, mes, float(dp["Valor"])))
            except (KeyError, ValueError, TypeError):
                continue
        return rows

    def _find_series(
        raw: list[dict],
        must_contain: list[str],
        must_not: list[str] | None = None,
    ) -> list[dict]:
        must_not = must_not or []
        matches = []
        for serie in raw:
            nombre = serie.get("Nombre", "").lower()
            if provincia.lower() not in nombre:
                continue
            if not all(t.lower() in nombre for t in must_contain):
                continue
            if any(t.lower() in nombre for t in must_not):
                continue
            matches.append(serie)
        return matches

    try:
        tabla_v = int(params.get("tabla_viajeros", tabla_viajeros_default))
        tabla_p = int(params.get("tabla_pernoctaciones", tabla_viajeros_default))
        total = 0

        raw_v = _fetch_tabla(tabla_v)
        series_v = _find_series(raw_v, must_contain=["viajero"])
        if not series_v:
            if verbose:
                print(f"  [ine_eoh] viajeros: ninguna serie para '{provincia}' en tabla {tabla_v}")
        else:
            agg: dict[tuple[int, int], float] = {}
            for s in series_v:
                for yr, mes, val in _parse_serie_mensual(s):
                    agg[(yr, mes)] = agg.get((yr, mes), 0.0) + val
            series_total = _find_series(raw_v, must_contain=["viajero", "total"])
            if series_total:
                agg = {}
                for s in series_total:
                    for yr, mes, val in _parse_serie_mensual(s):
                        agg[(yr, mes)] = val

            ensure_feature_registry(
                feature_key_viajeros,
                "ine_eoh",
                "turismo",
                f"Viajeros hoteleros estimados — {provincia} (INE EOH)",
            )
            for (yr, mes), val in sorted(agg.items()):
                total += write_month_uniform(yr, mes, val, loc_uuid, feature_key_viajeros, verbose)

        raw_p = _fetch_tabla(tabla_p) if tabla_p != tabla_v else raw_v
        series_p = _find_series(raw_p, must_contain=["pernoctaci"])
        series_p_total = _find_series(raw_p, must_contain=["pernoctaci", "total"])
        if series_p_total:
            series_p = series_p_total

        if not series_p:
            if verbose:
                print(
                    f"  [ine_eoh] pernoctaciones: ninguna serie para '{provincia}' en tabla {tabla_p}"
                )
        else:
            agg_p: dict[tuple[int, int], float] = {}
            for s in series_p:
                for yr, mes, val in _parse_serie_mensual(s):
                    agg_p[(yr, mes)] = val

            ensure_feature_registry(
                feature_key_pernoctaciones,
                "ine_eoh",
                "turismo",
                f"Pernoctaciones hoteleras estimadas — {provincia} (INE EOH)",
            )
            for (yr, mes), val in sorted(agg_p.items()):
                total += write_month_uniform(
                    yr, mes, val, loc_uuid, feature_key_pernoctaciones, verbose
                )

        write_sync_marker(loc_uuid, "ine_eoh")
        return total
    except Exception as e:
        if verbose:
            print(f"  [ine_eoh] {loc_uuid} ERROR — {e}")
        return 0


def _handler_esri_places(
    loc_uuid: str,
    params: dict,
    max_age_hours: float,
    verbose: bool,
) -> int:
    """Esri ArcGIS Places API — POIs por ubicacion."""
    if is_fresh(loc_uuid, "esri_places", max_age_hours):
        return 0

    token = os.environ.get("ESRI_KEY", "")
    if not token:
        if verbose:
            print(f"  [esri_places] {loc_uuid}: ESRI_KEY no encontrado, omitido")
        return 0

    try:
        n = sync_esri_places_location(loc_uuid, params, verbose)
        write_sync_marker(loc_uuid, "esri_places")
        return n
    except Exception as e:
        if verbose:
            print(f"  [esri_places] {loc_uuid} ERROR — {e}")
        return 0


# ── Dispatch table ────────────────────────────────────────────────────────────

_HANDLERS: dict[str, Callable] = {
    "metro_madrid": _handler_metro_madrid,
    "puertos_estado": _handler_puertos_estado,
    "ine_eoh": _handler_ine_eoh,
    "esri_places": _handler_esri_places,
}


# ── Funcion publica sync_all ──────────────────────────────────────────────────


def sync_all(
    location_uuid: str | None = None,
    max_age_hours: float = 168,
    verbose: bool = True,
) -> dict[str, dict[str, int]]:
    """
    Para cada source en _HANDLERS:
      1. Lee get_configured_locations(source) → ubicaciones con config activa
      2. Para cada ubicacion: ejecuta el handler si no es fresh
      3. Escribe sync marker (dentro de cada handler)
    Retorna {source: {location_uuid: n_rows}}.
    """

    def log(msg: str) -> None:
        if verbose:
            print(msg)

    log(
        f"\n  sync_mensual/sync_all — {len(_HANDLERS)} source(s): "
        f"{', '.join(sorted(_HANDLERS))}"
    )

    results: dict[str, dict[str, int]] = {}

    for name, handler in sorted(_HANDLERS.items()):
        configured = get_configured_locations(name)
        if location_uuid is not None:
            configured = [(lu, p) for lu, p in configured if lu == location_uuid]

        src_stats: dict[str, int] = {}
        for lu, params in configured:
            try:
                n = handler(lu, params, max_age_hours, verbose)
                src_stats[lu] = n
            except Exception as e:
                log(f"  [!] {name}: {lu} ERROR — {e}")
                src_stats[lu] = 0

        results[name] = src_stats
        total = sum(src_stats.values()) if src_stats else 0
        log(f"  [{name}] {total} filas escritas")

    return results


# ── Catalog (para context_scout.py) ──────────────────────────────────────────


def cargar_catalog(pais: str = "") -> list[dict]:
    """Devuelve entradas de catálogo de source_registry para context_scout.py."""
    from src.db.store import get_conn

    if pais:
        rows = (
            get_conn()
            .execute(
                "SELECT source, categoria, periodicidad, descripcion, url_referencia, "
                "cobertura_desde, latencia_dias, paises, params_schema, params_ejemplo, config "
                "FROM source_registry WHERE activo = TRUE "
                "AND (paises = '[]'::jsonb OR paises @> %s::jsonb)",
                [f'["{pais}"]'],
            )
            .fetchall()
        )
    else:
        rows = (
            get_conn()
            .execute(
                "SELECT source, categoria, periodicidad, descripcion, url_referencia, "
                "cobertura_desde, latencia_dias, paises, params_schema, params_ejemplo, config "
                "FROM source_registry WHERE activo = TRUE",
            )
            .fetchall()
        )
    cols = [
        "source",
        "categoria",
        "periodicidad",
        "descripcion",
        "url_referencia",
        "cobertura_desde",
        "latencia_dias",
        "paises",
        "params_schema",
        "params_ejemplo",
        "config",
    ]
    return [dict(zip(cols, r)) for r in rows]


# ── sync_esri_places_location (publica, para admin_pois.py) ──────────────────


def sync_esri_places_location(
    location_uuid: str,
    params: dict | None = None,
    verbose: bool = True,
) -> int:
    """
    Llama a Esri Places, upsertea resultados en location_pois.
    Expuesta publicamente para uso desde src/callbacks/admin_pois.py.
    Devuelve numero de POIs procesados.
    """
    token = os.environ.get("ESRI_KEY", "")
    if not token:
        if verbose:
            print(f"  [esri_places] {location_uuid}: ESRI_KEY no encontrado, omitido")
        return 0

    from src.db.queries import upsert_poi
    from src.db.store import get_conn

    params = params or {}

    cfg = get_source_config("esri_places", params)
    base_url = cfg["base_url"]
    near_point_url = base_url + "/places/near-point"
    radio_m_default = int(cfg.get("radio_m", 1200))
    page_size = int(cfg.get("page_size", 20))
    max_cat_per_call = int(cfg.get("max_category_ids_per_call", 10))
    cat_map_raw = cfg.get("categorias", {})
    cat_map_default: dict[str, tuple[str, str]] = {k: tuple(v) for k, v in cat_map_raw.items()}  # type: ignore[misc]
    valores_cat: dict[str, float] = cfg.get("valores_categoria", {"otro": 0.50})

    def _get_loc_coords():
        row = (
            get_conn()
            .execute(
                "SELECT lat, lon FROM dim_ubicaciones "
                "WHERE location_uuid = ? AND lat IS NOT NULL",
                [location_uuid],
            )
            .fetchone()
        )
        return (row[0], row[1]) if row else None

    def _get_org_uuid():
        row = (
            get_conn()
            .execute(
                "SELECT org_uuid FROM dim_ubicaciones WHERE location_uuid = ?",
                [location_uuid],
            )
            .fetchone()
        )
        return row[0] if row else None

    def _call_near_point(
        lat: float,
        lon: float,
        radio_m: int,
        category_ids: list[str],
        page_token: str | None = None,
    ) -> dict:
        p: dict = {
            "y": lat,
            "x": lon,
            "radius": radio_m,
            "categoryIds": ",".join(category_ids),
            "pageSize": page_size,
            "f": "json",
            "token": token,
        }
        if page_token:
            p["pageToken"] = page_token
        url = near_point_url + "?" + urllib.parse.urlencode(p)
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read())

    def _fetch_page_batch(
        lat: float,
        lon: float,
        radio_m: int,
        category_ids: list[str],
        max_resultados: int,
    ) -> list[dict]:
        results = []
        page_token = None
        while len(results) < max_resultados:
            try:
                data = _call_near_point(lat, lon, radio_m, category_ids, page_token)
            except Exception as e:
                raise RuntimeError(f"Error llamando Places API: {e}") from e
            if "error" in data:
                raise RuntimeError(f"Esri Places error: {data['error']}")
            batch = data.get("results", [])
            results.extend(batch)
            pagination = data.get("pagination", {})
            next_url = pagination.get("nextUrl")
            if not next_url or not batch:
                break
            parsed = urllib.parse.urlparse(next_url)
            qs = urllib.parse.parse_qs(parsed.query)
            page_token = qs.get("pageToken", [None])[0]
            if not page_token:
                break
        return results

    def _fetch_all_places(
        lat: float, lon: float, radio_m: int, category_ids: list[str], max_resultados: int
    ) -> list[dict]:
        seen: set[str] = set()
        results: list[dict] = []
        batches = [
            category_ids[i : i + max_cat_per_call]
            for i in range(0, len(category_ids), max_cat_per_call)
        ]
        for batch in batches:
            for place in _fetch_page_batch(lat, lon, radio_m, batch, max_resultados):
                pid = place.get("placeId", "")
                if pid and pid in seen:
                    continue
                if pid:
                    seen.add(pid)
                results.append(place)
                if len(results) >= max_resultados:
                    return results
        return results

    def _place_to_poi(place: dict, cat_map: dict) -> dict | None:
        name = (place.get("name") or "").strip()
        if not name:
            return None
        loc = place.get("location", {})
        lon = loc.get("x")
        lat = loc.get("y")
        if lat is None or lon is None:
            return None
        cats = place.get("categories", [])
        categoria = "otro"
        valor = valores_cat.get("otro", 0.50)
        for c in cats:
            cid = str(c.get("categoryId", ""))
            if cid in cat_map:
                categoria, _ = cat_map[cid]
                valor = valores_cat.get(categoria, 0.5)
                break
        cat_labels = [c.get("label", "") for c in cats if c.get("label")]
        detalle = " · ".join(cat_labels[:3]) if cat_labels else None
        return {
            "nombre": name,
            "lat": lat,
            "lon": lon,
            "categoria": categoria,
            "valor_relativo": valor,
            "detalle": detalle,
        }

    coords = _get_loc_coords()
    if not coords:
        if verbose:
            print(f"  [esri_places] {location_uuid}: sin coordenadas, omitido")
        return 0

    org_uuid = _get_org_uuid()
    if not org_uuid:
        return 0

    lat, lon = coords
    radio_m = int(params.get("radio_m", radio_m_default))
    max_res = int(params.get("max_resultados", 200))

    cat_ids_cfg = params.get("categorias")
    if cat_ids_cfg:
        cat_map: dict[str, tuple[str, str]] = {
            cid: cat_map_default.get(cid, ("otro", cid)) for cid in cat_ids_cfg
        }
    else:
        cat_map = cat_map_default.copy()

    try:
        places = _fetch_all_places(lat, lon, radio_m, list(cat_map.keys()), max_res)
    except RuntimeError as e:
        if verbose:
            print(f"  [esri_places] {location_uuid} ERROR — {e}")
        return 0

    n = 0
    for place in places:
        poi = _place_to_poi(place, cat_map)
        if poi is None:
            continue
        try:
            upsert_poi(
                location_uuid=location_uuid,
                org_uuid=org_uuid,
                fuente="esri_places",
                **poi,
            )
            n += 1
        except Exception as e:
            if verbose:
                print(f"  [esri_places] upsert error '{poi['nombre']}': {e}")

    if verbose:
        print(
            f"  [esri_places] {location_uuid}: {n} POIs sincronizados ({len(places)} encontrados)"
        )
    return n


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Sincronizacion mensual completa (todos los sources configurados)"
    )
    parser.add_argument("--location", metavar="UUID")
    parser.add_argument(
        "--solo",
        default=None,
        metavar="SOURCE[,SOURCE]",
        help="Ejecutar solo este source (ej: metro_madrid)",
    )
    parser.add_argument("--max-age", type=float, default=168, metavar="HORAS")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args()

    effective_handlers = _HANDLERS
    if args.solo:
        solo_set = set(args.solo.split(","))
        effective_handlers = {k: v for k, v in _HANDLERS.items() if k in solo_set}

    _original_handlers = _HANDLERS.copy()
    _HANDLERS.clear()
    _HANDLERS.update(effective_handlers)

    sync_all(
        location_uuid=args.location,
        max_age_hours=0 if args.force else args.max_age,
        verbose=not args.quiet,
    )

    _HANDLERS.clear()
    _HANDLERS.update(_original_handlers)
