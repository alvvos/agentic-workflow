"""
Metro de Madrid — validaciones mensuales por estación (CRTM / Metro Madrid).

Fuente: Portal open data de Metro de Madrid.
  Informe anual: https://www.metromadrid.es/en/metro-de-madrid/statistics
  Dataset CRTM: https://datos.crtm.es (validaciones por mes/estación/línea)

Formato del Excel publicado por Metro Madrid:
  Hoja "Viajeros" (u hoja por año): filas = estaciones, columnas = meses (1-12) + Total.
  Columna A: nombre de estación.  Columnas B-M: Ene-Dic (validaciones enteras).

Convención de feature_key: afluencia_metro_{slug}
  Ej: afluencia_metro_gran_via, afluencia_metro_callao, afluencia_metro_sol

Configuración en location_source_config (source = 'metro_madrid'):
  {
    "estaciones": [
      {"nombre": "Gran Vía",      "slug": "gran_via"},
      {"nombre": "Callao",        "slug": "callao"},
      {"nombre": "Sol",           "slug": "sol"},
      {"nombre": "Santo Domingo", "slug": "santo_domingo"}
    ],
    "anyo_url": "https://www.metromadrid.es/export/sites/metro/comun/documentos/viajeros/Estadistica_{year}.xlsx"
  }

El campo "anyo_url" es opcional — si no se especifica se usa _DEFAULT_URL_PATTERN.

Escribe en store_features_ext con periodicidad diaria (valor mensual distribuido
uniformemente en los días del mes).

CLI:
    python -m src.data_ingestion.mensual.metro_madrid
    python -m src.data_ingestion.mensual.metro_madrid --force
    python -m src.data_ingestion.mensual.metro_madrid --list-stations
"""

from __future__ import annotations

import calendar
import io
import json
import sys
import warnings
from datetime import date
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from src.data_ingestion.prefetch._common import is_fresh, write_sync_marker
from src.db.store import get_conn

# ── Declaraciones de paquete mensual ─────────────────────────────────────────

SOURCE = "metro_madrid"

CATALOG_PAISES = ["ES"]

CATALOG_ENTRY = {
    "feature_key_template": "afluencia_metro_{slug}",
    "source": SOURCE,
    "categoria": "movilidad",
    "periodicidad": "mensual",
    "descripcion": (
        "Validaciones mensuales por estación de metro (Metro de Madrid / CRTM). "
        "Mide el número de accesos validados en cada estación dentro de la isócrona "
        "de la ubicación. Proxy directo del volumen de peatones que pasan por el área."
    ),
    "url_referencia": "https://www.metromadrid.es/en/metro-de-madrid/statistics",
    "granularidad": "estación (mensual, distribuida en días)",
    "cobertura_desde": "2016-01",
    "latencia_dias": 45,
    "notas_tecnicas": (
        "Requiere configurar 'estaciones' en location_source_config. "
        "Cada entrada necesita 'nombre' (exacto en el Excel) y 'slug' para el feature_key. "
        "Si el URL del Excel cambia, actualiza 'anyo_url' en location_source_config."
    ),
    "params_schema": (
        "{'estaciones': [{'nombre': '<nombre exacto en el Excel de Metro Madrid>', "
        "'slug': '<snake_case del nombre>'}], "
        "'anyo_url': '<opcional — URL pattern con {year}>'}. "
        "Incluir las 2-4 estaciones de Metro de Madrid a ≤800 m de las coordenadas."
    ),
    "params_ejemplo": {
        "estaciones": [
            {"nombre": "Gran Vía", "slug": "gran_via"},
            {"nombre": "Callao", "slug": "callao"},
        ]
    },
}

_TIMEOUT = 30
_DEFAULT_URL_PATTERN = (
    "https://www.metromadrid.es/export/sites/metro/comun/documentos/viajeros/"
    "Estadistica_{year}.xlsx"
)

_MESES_COL = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]


# ── Config desde location_source_config ──────────────────────────────────────


def _get_configured_locations() -> list[tuple[str, list[dict], str]]:
    """Returns [(location_uuid, estaciones, url_pattern), ...]."""
    rows = (
        get_conn()
        .execute(
            "SELECT location_uuid, params "
            "FROM location_source_config "
            "WHERE source = 'metro_madrid' AND activo = TRUE"
        )
        .fetchall()
    )
    result = []
    for loc_uuid, params_raw in rows:
        params = params_raw if isinstance(params_raw, dict) else json.loads(params_raw or "{}")
        estaciones = params.get("estaciones", [])
        url_pattern = params.get("anyo_url", _DEFAULT_URL_PATTERN)
        if estaciones:
            result.append((loc_uuid, estaciones, url_pattern))
    return result


# ── Descarga y parsing ────────────────────────────────────────────────────────


def _download_excel(year: int, url_pattern: str) -> bytes | None:
    url = url_pattern.format(year=year)
    try:
        r = requests.get(url, timeout=_TIMEOUT)
        if r.status_code in (404, 403):
            return None
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def _parse_excel_year(
    xlsx_bytes: bytes,
    estaciones: list[dict],
    year: int,
) -> dict[str, dict[int, int]]:
    """
    Parsea el Excel de validaciones anuales de Metro Madrid.
    Devuelve {slug: {mes: total_validaciones}}.

    Formato esperado:
      - Filas 1-N: estaciones (col A = nombre, cols B-M = meses Ene-Dic).
      - Podría haber varias hojas por año; probamos la que contenga más datos.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl es necesario: pip install openpyxl") from exc

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    # Intentar hoja con nombre del año o primera hoja con datos numéricos
    target_sheet = None
    for sh_name in wb.sheetnames:
        if str(year) in sh_name:
            target_sheet = wb[sh_name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    ws = target_sheet

    # Detectar cabecera de meses — buscar fila donde col A está vacía y cols B+ son
    # nombres de meses o números
    header_row: int | None = None
    for r in range(1, min(10, ws.max_row + 1)):
        vals_b_m = [ws.cell(r, c).value for c in range(2, 14)]
        non_null = [v for v in vals_b_m if v is not None]
        if len(non_null) >= 6:
            # Comprobar si parecen meses (texto o número 1-12)
            first = non_null[0]
            if isinstance(first, str) and any(m in str(first).lower() for m in _MESES_COL):
                header_row = r
                break
            if isinstance(first, (int, float)) and 1 <= int(first) <= 12:
                header_row = r
                break

    # Determinar qué columnas corresponden a cada mes
    if header_row is not None:
        col_mes: dict[int, int] = {}  # {mes_num: col_idx}
        for c in range(2, 14):
            v = ws.cell(header_row, c).value
            if v is None:
                continue
            if isinstance(v, str):
                for i, mn in enumerate(_MESES_COL, 1):
                    if mn in v.lower():
                        col_mes[i] = c
                        break
            elif isinstance(v, (int, float)) and 1 <= int(v) <= 12:
                col_mes[int(v)] = c
    else:
        # Asumir columnas B-M = meses 1-12
        col_mes = {m: m + 1 for m in range(1, 13)}
        header_row = 1

    # Buscar estaciones por nombre en col A
    nombre_to_slug = {e["nombre"].strip().lower(): e["slug"] for e in estaciones}
    result: dict[str, dict[int, int]] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        cell_a = ws.cell(r, 1).value
        if cell_a is None:
            continue
        nombre_celda = str(cell_a).strip().lower()
        for nombre_cfg, slug in nombre_to_slug.items():
            if nombre_cfg in nombre_celda or nombre_celda in nombre_cfg:
                mes_vals: dict[int, int] = {}
                for mes, col in col_mes.items():
                    v = ws.cell(r, col).value
                    try:
                        mes_vals[mes] = int(float(str(v).replace(",", "").strip()))
                    except (TypeError, ValueError):
                        pass
                if mes_vals:
                    result[slug] = mes_vals
                break

    return result


# ── Escritura en DB ───────────────────────────────────────────────────────────


def _write_month(
    year: int,
    month: int,
    total_validaciones: int,
    location_uuid: str,
    slug: str,
    verbose: bool = False,
) -> int:
    if total_validaciones <= 0:
        return 0
    today = date.today()
    last_day = calendar.monthrange(year, month)[1]
    if date(year, month, last_day) >= today:
        return 0

    feature_key = f"afluencia_metro_{slug}"
    val_per_day = total_validaciones / last_day
    rows = [
        (str(date(year, month, d)), location_uuid, feature_key, val_per_day)
        for d in range(1, last_day + 1)
    ]
    get_conn().executemany(
        "INSERT INTO store_features_ext (fecha, location_uuid, feature_key, value) "
        "VALUES (?,?,?,?) "
        "ON CONFLICT (fecha, location_uuid, feature_key) "
        "DO UPDATE SET value = excluded.value, ingested_at = NOW()",
        rows,
    )
    if verbose:
        print(
            f"  [metro_madrid] {slug} {month:02d}/{year}: "
            f"{total_validaciones:,} validaciones → {val_per_day:.0f}/día"
        )
    return len(rows)


# ── Asegurar feature_registry ─────────────────────────────────────────────────


def _ensure_feature_registry(slug: str) -> None:
    fk = f"afluencia_metro_{slug}"
    conn = get_conn()
    exists = conn.execute("SELECT 1 FROM feature_registry WHERE feature_key = ?", [fk]).fetchone()
    if not exists:
        conn.execute(
            "INSERT INTO feature_registry "
            "(feature_key, source, descripcion, categoria, periodicidad, activo) "
            "VALUES (?,?,?,?,?,TRUE) ON CONFLICT (feature_key) DO NOTHING",
            [
                fk,
                SOURCE,
                f"Validaciones mensuales estación de metro — {slug.replace('_', ' ').title()}",
                "movilidad",
                "mensual",
            ],
        )


# ── Sync principal ────────────────────────────────────────────────────────────


def sync_year(
    year: int,
    location_uuid: str,
    estaciones: list[dict],
    url_pattern: str,
    verbose: bool = True,
) -> int:
    raw = _download_excel(year, url_pattern)
    if raw is None:
        if verbose:
            print(f"  [metro_madrid] {year}: descarga fallida o no disponible")
        return 0

    parsed = _parse_excel_year(raw, estaciones, year)
    if not parsed:
        if verbose:
            print(f"  [metro_madrid] {year}: ninguna estación encontrada en el Excel")
        return 0

    total = 0
    for slug, mes_vals in parsed.items():
        _ensure_feature_registry(slug)
        for mes, validaciones in mes_vals.items():
            total += _write_month(year, mes, validaciones, location_uuid, slug, verbose)

    return total


def sync(jobs: list, fecha: date) -> int:
    """Interfaz estándar para sync_mensual."""
    locations = _get_configured_locations()
    total = 0
    for loc_uuid, estaciones, url_pattern in locations:
        total += sync_year(fecha.year - 1, loc_uuid, estaciones, url_pattern)
        total += sync_year(fecha.year, loc_uuid, estaciones, url_pattern)
    return total


def run(
    location_uuid: str | None = None,
    max_age_hours: float = 168,
    verbose: bool = True,
) -> dict[str, int]:
    """Interfaz nightly (prefetch). Por defecto semanal — dato es mensual."""
    locations = _get_configured_locations()
    if location_uuid is not None:
        locations = [(lu, e, u) for lu, e, u in locations if lu == location_uuid]

    result: dict[str, int] = {}
    today = date.today()

    for loc_uuid, estaciones, url_pattern in locations:
        if is_fresh(loc_uuid, SOURCE, max_age_hours):
            result[loc_uuid] = 0
            continue
        try:
            n = sync_year(today.year - 1, loc_uuid, estaciones, url_pattern, verbose)
            n += sync_year(today.year, loc_uuid, estaciones, url_pattern, verbose)
            write_sync_marker(loc_uuid, SOURCE)
            result[loc_uuid] = n
        except Exception as e:
            if verbose:
                print(f"  [metro_madrid] {loc_uuid} ERROR — {e}")
            result[loc_uuid] = 0

    return result


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Ingestor validaciones Metro de Madrid")
    parser.add_argument("--force", action="store_true", help="Ignora caché de frescura")
    parser.add_argument(
        "--list-stations", action="store_true", help="Imprime estaciones configuradas y sale"
    )
    parser.add_argument("--year", type=int, default=None, help="Año concreto a ingestar")
    args = parser.parse_args()

    locations = _get_configured_locations()
    if not locations:
        print("[metro_madrid] No hay ubicaciones configuradas en location_source_config.")
        sys.exit(0)

    if args.list_stations:
        for loc_uuid, estaciones, _ in locations:
            print(f"  {loc_uuid}: {[e['nombre'] for e in estaciones]}")
        sys.exit(0)

    today = date.today()
    for loc_uuid, estaciones, url_pattern in locations:
        years = [args.year] if args.year else [today.year - 1, today.year]
        for yr in years:
            sync_year(yr, loc_uuid, estaciones, url_pattern, verbose=True)
