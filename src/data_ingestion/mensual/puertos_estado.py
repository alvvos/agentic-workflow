"""
Puertos del Estado — estadística mensual oficial de pasajeros de crucero.

Fuente: https://www.puertos.es/en/data/statistics/monthly
XLSX publicado ~día 22-25 de cada mes con 4-6 semanas de retraso respecto al cierre.
Acceso mediante GET con parámetro date_value (date_value = 2027 - año).

Estructura del XLSX (hoja "Pasajeros crucero", 36 filas x 14 cols):
  R5:  ["Autoridad Portuaria", "<Mes> ", "", "Acumulado desde Enero", ...]
  R6:  ["", año_anterior, año_actual, año_anterior, año_actual, "Var. (%)", ...]
  R7+: [nombre_AP, pax_mes_ant, pax_mes_act, pax_ytd_ant, pax_ytd_act, var]

Cada fichero contiene datos del mes propio (cols B/C) y su acumulado YTD (cols D/E),
más el mismo dato del año anterior en columna B/D.
Esto permite extraer dos años de datos descargando solo los ficheros del año actual.

Escribe en store_features_ext:
  n_pasajeros_crucero_oficial — total mensual distribuido uniformemente en días

Las ubicaciones activas y su AP se leen de location_source_config (source='puertos_estado').
El campo params debe contener: {"port_authority": "<nombre exacto en la hoja XLSX>"}

CLI:
    python -m src.data_ingestion.mensual.puertos_estado
    python -m src.data_ingestion.mensual.puertos_estado --desde 2024-01
    python -m src.data_ingestion.mensual.puertos_estado --force
"""

from __future__ import annotations

import calendar
import io
import re
import sys
import warnings
from datetime import date, datetime
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from src.data_ingestion.prefetch._common import is_fresh, write_sync_marker
from src.db.store import get_conn

# ── Declaraciones de paquete mensual ─────────────────────────────────────────

SOURCE = "puertos_estado"

CATALOG_PAISES = ["ES"]

CATALOG_ENTRY = {
    "feature_key_template": "n_pasajeros_crucero_oficial",
    "source": "puertos_estado",
    "categoria": "turismo",
    "periodicidad": "mensual",
    "descripcion": (
        "Pasajeros de crucero — Puertos del Estado (estadística oficial). "
        "Total mensual de pasajeros embarcados y desembarcados en el puerto. "
        "Cuenta personas reales que pasan físicamente por el puerto cada mes. "
        "Proxy directo del flujo de turistas de crucero hacia el centro de la ciudad."
    ),
    "url_referencia": "https://www.puertos.es/en/data/statistics/monthly",
    "url_descarga": "https://www.puertos.es/en/data/statistics/monthly",
    "granularidad": "puerto (ciudad)",
    "cobertura_desde": "2012-01",
    "latencia_dias": 25,
    "notas_tecnicas": (
        "XLSX publicado ~día 22-25 del mes siguiente. Sin autenticación. "
        "Solo incluir si la ubicación está en una ciudad con puerto comercial activo "
        "de cruceros (Málaga, Barcelona, Palma, Sevilla, Las Palmas, etc.). "
        "Requiere configurar port_authority en location_source_config con el nombre "
        "exacto de la Autoridad Portuaria tal como aparece en la hoja XLSX."
    ),
}

_SOURCE = SOURCE
_FEATURE_KEY = "n_pasajeros_crucero_oficial"
_TIMEOUT = 30
_BASE_URL = "https://www.puertos.es"
_LISTING_URL = _BASE_URL + "/en/data/statistics/monthly"
_SHEET = "Pasajeros crucero"

# date_value parameter: 2027 - year (2026→1, 2025→2, 2024→3, ...)
_MIN_YEAR = 2012  # primer año disponible en el portal


# ── Config desde location_source_config ──────────────────────────────────────


def _get_configured_locations() -> list[tuple[str, str]]:
    """Returns [(location_uuid, port_authority), ...] for all active puertos_estado configs."""
    rows = (
        get_conn()
        .execute(
            "SELECT location_uuid, params->>'port_authority' "
            "FROM location_source_config "
            "WHERE source = 'puertos_estado' AND activo = TRUE"
        )
        .fetchall()
    )
    return [(r[0], r[1]) for r in rows if r[1]]


# ── Descubrimiento de IDs ──────────────────────────────────────────────────────


def _fetch_listing_ids(year: int) -> dict[int, int]:
    """
    Raspa la página de listado para el año dado y devuelve {mes: file_id_xlsx}.
    Usa el parámetro GET date_value = 2027 - year.
    Los IDs del XLSX están en las posiciones pares de la lista de todos los links.
    """
    date_value = 2027 - year
    if date_value < 1:
        return {}
    try:
        r = requests.get(_LISTING_URL, params={"date_value": date_value}, timeout=_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return {}

    all_ids = re.findall(r"/file-download/download/public/(\d+)", r.text)
    pdf_months = re.findall(r"CuadrosResumen_\d{4}_(\d{2})\.pdf", r.text)

    # Los IDs aparecen en pares (XLSX, PDF) en el mismo orden que los meses PDF
    xlsx_ids = all_ids[::2]

    month_ids: dict[int, int] = {}
    for fid, mo_str in zip(xlsx_ids, pdf_months):
        mo = int(mo_str)
        if mo not in month_ids:
            month_ids[mo] = int(fid)

    return month_ids


# ── Descarga y parsing ────────────────────────────────────────────────────────


def _download_xlsx(file_id: int) -> bytes | None:
    url = f"{_BASE_URL}/file-download/download/public/{file_id}"
    try:
        r = requests.get(url, timeout=_TIMEOUT)
        if r.status_code in (404, 403):
            return None
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def parse_xlsx(xlsx_bytes: bytes, port_authority: str) -> dict | None:
    """
    Parsea la hoja "Pasajeros crucero" y devuelve:
    {
        'month'    : int,  # número de mes (1-12)
        'year_act' : int,  # año del dato actual (col C)
        'year_prev': int,  # año del dato anterior (col B)
        'pax_act'  : int,  # pasajeros mes año actual para port_authority
        'pax_prev' : int,  # pasajeros mes año anterior para port_authority
    }
    Retorna None si el fichero no tiene el formato esperado o no encuentra la AP.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl es necesario: pip install openpyxl") from exc

    _MESES_ES = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    if _SHEET not in wb.sheetnames:
        return None
    ws = wb[_SHEET]

    # R5: cabecera con nombre del mes (col B)
    month_num: int | None = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(5, col).value
        if v and isinstance(v, str):
            for name, num in _MESES_ES.items():
                if name in v.lower():
                    month_num = num
                    break
        if month_num:
            break

    # R6: años en columnas B (previo) y C (actual)
    year_prev: int | None = None
    year_act: int | None = None
    for col in range(2, 6):
        v = ws.cell(6, col).value
        if v and isinstance(v, (int, float)):
            yr = int(v)
            if 2010 < yr < 2100:
                if year_prev is None:
                    year_prev = yr
                elif year_act is None:
                    year_act = yr
                    break

    if month_num is None or year_act is None or year_prev is None:
        return None

    # Buscar fila de la AP (escaneo completo por robustez)
    ap_row: int | None = None
    for r in range(7, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and port_authority.lower() == str(v).strip().lower():
            ap_row = r
            break

    if ap_row is None:
        return None

    def _int(v) -> int:
        try:
            return int(float(str(v).replace(",", "").strip()))
        except (TypeError, ValueError):
            return 0

    return {
        "month": month_num,
        "year_act": year_act,
        "year_prev": year_prev,
        "pax_act": _int(ws.cell(ap_row, 3).value),  # col C = año actual
        "pax_prev": _int(ws.cell(ap_row, 2).value),  # col B = año anterior
    }


# ── Escritura en DB ───────────────────────────────────────────────────────────


def _write_month(
    year: int, month: int, total_pax: int, location_uuid: str, verbose: bool = False
) -> int:
    """
    Distribuye el total mensual uniformemente en los días del mes y escribe
    en store_features_ext. Solo escribe meses ya cerrados (último día < hoy).
    Idempotente. Devuelve número de filas escritas.
    """
    today = date.today()
    last_day = calendar.monthrange(year, month)[1]
    if date(year, month, last_day) >= today:
        return 0

    pax_per_day = total_pax / last_day
    rows = [
        (str(date(year, month, d)), location_uuid, _FEATURE_KEY, pax_per_day)
        for d in range(1, last_day + 1)
    ]

    conn = get_conn()
    conn.executemany(
        "INSERT INTO store_features_ext (fecha, location_uuid, feature_key, value) "
        "VALUES (?,?,?,?) "
        "ON CONFLICT (fecha, location_uuid, feature_key) "
        "DO UPDATE SET value = excluded.value, ingested_at = NOW()",
        rows,
    )
    return len(rows)


# ── Sync por año ──────────────────────────────────────────────────────────────


def sync_year(year: int, location_uuid: str, port_authority: str, verbose: bool = True) -> int:
    """
    Descarga todos los ficheros disponibles para el año y persiste datos de la AP.
    Cada fichero aporta datos del mes propio (year_act) y del mismo mes del año
    anterior (year_prev), maximizando cobertura histórica con el mínimo de descargas.
    Retorna total de días escritos.
    """
    ids = _fetch_listing_ids(year)
    if not ids:
        if verbose:
            print(f"  [puertos_estado] {port_authority} {year}: no se encontraron ficheros")
        return 0

    total = 0
    for mes, fid in sorted(ids.items()):
        raw = _download_xlsx(fid)
        if raw is None:
            if verbose:
                print(
                    f"  [puertos_estado] {port_authority} {mes:02d}/{year}: "
                    f"descarga fallida (ID {fid})"
                )
            continue

        parsed = parse_xlsx(raw, port_authority)
        if parsed is None:
            if verbose:
                print(
                    f"  [puertos_estado] {port_authority} {mes:02d}/{year}: "
                    f"formato inesperado o AP no encontrada (ID {fid})"
                )
            continue

        n = _write_month(parsed["year_act"], parsed["month"], parsed["pax_act"], location_uuid)
        if n > 0:
            total += n
            if verbose:
                print(
                    f"  [puertos_estado] {port_authority} {parsed['month']:02d}/{parsed['year_act']}: "
                    f"{parsed['pax_act']:,} pax"
                )

        # El fichero también contiene el mismo mes del año anterior
        n_prev = _write_month(
            parsed["year_prev"], parsed["month"], parsed["pax_prev"], location_uuid
        )
        if n_prev > 0:
            total += n_prev
            if verbose:
                print(
                    f"  [puertos_estado] {port_authority} {parsed['month']:02d}/{parsed['year_prev']}: "
                    f"{parsed['pax_prev']:,} pax (año ant.)"
                )

    return total


# ── Interfaz sync_mensual ─────────────────────────────────────────────────────


def sync(jobs: list, fecha: date) -> int:
    """Interfaz estándar para sync_mensual. Itera todas las ubicaciones configuradas."""
    locations = _get_configured_locations()
    total = 0
    for location_uuid, port_authority in locations:
        total += sync_year(fecha.year - 1, location_uuid, port_authority, verbose=True)
        total += sync_year(fecha.year, location_uuid, port_authority, verbose=True)
    return total


# ── run() — interfaz prefetch nightly ────────────────────────────────────────


def run(
    location_uuid: str | None = None,
    max_age_hours: float = 168,  # semanal por defecto: el dato es mensual
    verbose: bool = True,
) -> dict[str, int]:
    locations = _get_configured_locations()
    if location_uuid is not None:
        locations = [(lu, pa) for lu, pa in locations if lu == location_uuid]

    result: dict[str, int] = {}
    today = date.today()

    for loc_uuid, port_authority in locations:
        if is_fresh(loc_uuid, _SOURCE, max_age_hours):
            if verbose:
                print(f"  [puertos_estado] {port_authority} omitido (datos < {max_age_hours:.0f}h)")
            result[loc_uuid] = 0
            continue

        try:
            n = sync_year(today.year - 1, loc_uuid, port_authority, verbose=verbose)
            n += sync_year(today.year, loc_uuid, port_authority, verbose=verbose)
            write_sync_marker(loc_uuid, _SOURCE)
            if verbose:
                print(f"  [puertos_estado] {port_authority}: {n} días escritos")
            result[loc_uuid] = n
        except Exception as e:
            if verbose:
                print(f"  [puertos_estado] {port_authority} ERROR — {e}")
            result[loc_uuid] = 0

    return result


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Ingestor pasajeros crucero oficiales — Puertos del Estado"
    )
    parser.add_argument(
        "--desde",
        default=None,
        metavar="YYYY-MM",
        help="Año de inicio (solo se usa el año). Ej: 2024-01",
    )
    parser.add_argument(
        "--hasta",
        default=None,
        metavar="YYYY-MM",
        help="Año de fin (solo se usa el año). Ej: 2026-05",
    )
    parser.add_argument("--force", action="store_true", help="Ignora caché")
    parser.add_argument("--max-age", type=float, default=168, metavar="HORAS")
    args = parser.parse_args()

    def _ym(s: str) -> tuple[int, int]:
        d = datetime.strptime(s, "%Y-%m")
        return d.month, d.year

    today = date.today()
    locations = _get_configured_locations()
    if not locations:
        print("[puertos_estado] No hay ubicaciones configuradas en location_source_config.")
        sys.exit(0)

    if args.desde or args.hasta:
        d_from = _ym(args.desde) if args.desde else (1, today.year - 1)
        d_to = _ym(args.hasta) if args.hasta else (today.month, today.year)
        for yr in range(d_from[1], d_to[1] + 1):
            for loc_uuid, port_authority in locations:
                sync_year(yr, loc_uuid, port_authority)
    else:
        run(max_age_hours=0 if args.force else args.max_age)
