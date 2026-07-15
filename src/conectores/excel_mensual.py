"""
Conector para fuentes de datos mensuales distribuidas en Excel.

cfg["modo"] determina cómo se obtiene el fichero:
  "listado"  → scraping de página de listado para extraer file IDs (ej. Puertos del Estado)

Interfaz pública:
    TIPO = "excel_mensual"
    sync(ubicacion_id, cfg, verbose) -> int
"""

from __future__ import annotations

import io
import warnings
from datetime import date

import requests

from src.data_ingestion._common import write_month_uniform

TIPO = "excel_mensual"

_TIMEOUT = 30


# ── Modo "listado" — Puertos del Estado ──────────────────────────────────────


def _fetch_listing_ids(listing_url: str, year: int) -> dict[int, int]:
    import re

    date_value = 2027 - year
    if date_value < 1:
        return {}
    try:
        r = requests.get(listing_url, params={"date_value": date_value}, timeout=_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return {}
    all_ids = re.findall(r"/file-download/download/public/(\d+)", r.text)
    pdf_months = re.findall(r"CuadrosResumen_\d{4}_(\d{2})\.pdf", r.text)
    xlsx_ids = all_ids[::2]
    month_ids: dict[int, int] = {}
    for fid, mo_str in zip(xlsx_ids, pdf_months):
        mo = int(mo_str)
        if mo not in month_ids:
            month_ids[mo] = int(fid)
    return month_ids


def _download_xlsx_by_id(base_url: str, file_id: int) -> bytes | None:
    url = f"{base_url}/file-download/download/public/{file_id}"
    try:
        r = requests.get(url, timeout=_TIMEOUT)
        if r.status_code in (404, 403):
            return None
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def _parse_puertos_xlsx(xlsx_bytes: bytes, port_authority: str, sheet_name: str) -> dict | None:
    _MESES_ES = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl es necesario: pip install openpyxl") from exc

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    month_num: int | None = None
    for col in range(1, ws.max_column + 1):
        v = ws.cell(5, col).value
        if v and isinstance(v, str):
            for name, num in _MESES_ES.items():
                if name in v.lower():
                    month_num = num
                    break
        if month_num:
            break

    year_prev: int | None = None
    year_act: int | None = None
    for col in range(2, 6):
        v = ws.cell(6, col).value
        if v and isinstance(v, (int, float)):
            yr = int(v)
            if 2010 < yr < 2100:
                if year_prev is None:
                    year_prev = yr
                elif year_act is None:
                    year_act = yr
                    break

    if month_num is None or year_act is None or year_prev is None:
        return None

    ap_row: int | None = None
    for r in range(7, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and port_authority.lower() == str(v).strip().lower():
            ap_row = r
            break

    if ap_row is None:
        return None

    def _int(v) -> int:
        try:
            return int(float(str(v).replace(",", "").strip()))
        except (TypeError, ValueError):
            return 0

    return {
        "month": month_num,
        "year_act": year_act,
        "year_prev": year_prev,
        "pax_act": _int(ws.cell(ap_row, 3).value),
        "pax_prev": _int(ws.cell(ap_row, 2).value),
    }


def _sync_puertos_estado(ubicacion_id: str, cfg: dict, verbose: bool) -> int:
    port_authority = cfg.get("port_authority")
    if not port_authority:
        if verbose:
            print(f"  [excel_mensual/puertos] {ubicacion_id}: sin port_authority — omitido")
        return 0

    feature_key = cfg.get("feature_key", "n_pasajeros_crucero_oficial")
    listing_url = cfg.get("listing_url", "")
    base_url = listing_url.rsplit("/en/", 1)[0] if "/en/" in listing_url else listing_url
    sheet_name = cfg.get("hoja_excel", "Pasajeros crucero")

    def _sync_year(year: int) -> int:
        ids = _fetch_listing_ids(listing_url, year)
        if not ids:
            if verbose:
                print(
                    f"  [excel_mensual/puertos] {port_authority} {year}: no se encontraron ficheros"
                )
            return 0
        total = 0
        for mes, fid in sorted(ids.items()):
            raw = _download_xlsx_by_id(base_url, fid)
            if raw is None:
                if verbose:
                    print(
                        f"  [excel_mensual/puertos] {port_authority} {mes:02d}/{year}: "
                        f"descarga fallida (ID {fid})"
                    )
                continue
            parsed = _parse_puertos_xlsx(raw, port_authority, sheet_name)
            if parsed is None:
                if verbose:
                    print(
                        f"  [excel_mensual/puertos] {port_authority} {mes:02d}/{year}: "
                        f"formato inesperado o AP no encontrada (ID {fid})"
                    )
                continue
            n = write_month_uniform(
                parsed["year_act"], parsed["month"], parsed["pax_act"], ubicacion_id, feature_key
            )
            if n > 0:
                total += n
                if verbose:
                    print(
                        f"  [excel_mensual/puertos] {port_authority}"
                        f" {parsed['month']:02d}/{parsed['year_act']}: {parsed['pax_act']:,} pax"
                    )
            n_prev = write_month_uniform(
                parsed["year_prev"], parsed["month"], parsed["pax_prev"], ubicacion_id, feature_key
            )
            if n_prev > 0:
                total += n_prev
                if verbose:
                    print(
                        f"  [excel_mensual/puertos] {port_authority}"
                        f" {parsed['month']:02d}/{parsed['year_prev']}:"
                        f" {parsed['pax_prev']:,} pax (anyo ant.)"
                    )
        return total

    today = date.today()
    n = _sync_year(today.year - 1)
    n += _sync_year(today.year)
    if verbose:
        print(f"  [excel_mensual/puertos] {port_authority}: {n} dias escritos")
    return n


# ── Interfaz pública ──────────────────────────────────────────────────────────


def sync(ubicacion_id: str, cfg: dict, verbose: bool = True) -> int:
    """
    Descarga y persiste datos mensuales desde un Excel.

    ubicacion_id: UUID de la ubicación.
    cfg: config efectiva — debe contener "modo" ("url" o "listado").
    No llama a is_fresh() ni write_sync_marker() — los gestiona el orquestador.
    Devuelve el número de filas escritas.
    """
    modo = cfg.get("modo")
    if modo == "listado":
        try:
            return _sync_puertos_estado(ubicacion_id, cfg, verbose)
        except Exception as e:
            if verbose:
                print(f"  [excel_mensual/puertos] {ubicacion_id} ERROR — {e}")
            return 0
    else:
        if verbose:
            print(f"  [excel_mensual] modo '{modo}' desconocido para {ubicacion_id} — omitido")
        return 0
